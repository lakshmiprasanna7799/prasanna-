from openpyxl import Workbook
wb=Workbook()
ws=wb.active
ws['A1']="INVOICENO"
ws['B1']="STACKCODE"
ws['C1']="DESCRIPTION"
ws['D1']="QUALITY"
ws['E1']="UNITPRICE"

ws.append([2344,456,678,345,45])
#import datetime
#ws['A2']=datetime.datetime.now()
wb.save("sample1.xlsx")